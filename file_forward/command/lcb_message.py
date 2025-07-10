import os

import openpyxl

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from file_forward.configlib import add_config_option
from file_forward.configlib import engine_from_config
from file_forward.model import LCBMessageFilter
from file_forward.util import load_pyfile

HEADER = ['airline_code', 'date_of_origin', 'flight_number', 'departure_airport', 'destination_airport']

def add_parser(argument_parser):
    """
    Add lcb_message sub-commands to parser.
    """
    lcb_message_parser = argument_parser.add_parser(
        'lcb_message',
        help = 'LCB Message filter commands.',
    )
    add_config_option(lcb_message_parser)

    subparsers = lcb_message_parser.add_subparsers()

    # Export command
    export_command_parser = subparsers.add_parser(
        'export',
        help = 'Export LCB message filter data.',
    )
    add_config_option(export_command_parser)
    export_command_parser.add_argument(
        'output',
        help = 'Output filename.',
    )
    export_command_parser.add_argument(
        '--format',
        choices = ['excel'],
        default = 'excel',
        help = 'Output file format.',
    )
    export_command_parser.set_defaults(func=export_lcb_message_filter)

def get_last_column_letter(ws):
    return get_column_letter(ws.max_column)

def get_used_range(ws):
    """
    Return the Excel-style used range string of a worksheet, e.g., 'A1:D20'.
    """
    min_row = ws.min_row
    min_col = ws.min_column
    max_row = ws.max_row
    max_col = ws.max_column

    start_cell = f"{get_column_letter(min_col)}{min_row}"
    end_cell = f"{get_column_letter(max_col)}{max_row}"

    return f"{start_cell}:{end_cell}"

def export_lcb_message_filter(args):
    """
    Export LCB message filter data to a file.
    """
    if os.path.exists(args.output):
        raise FileExistsError(f'File exists {args.output}')

    appconfig = load_pyfile(args.config)
    engine = engine_from_config(appconfig)

    with Session(engine) as session:
        bold_font = Font(bold=True)

        wb = openpyxl.Workbook()
        ws = wb.active

        # Freeze first row.
        ws.freeze_panes = 'A2'

        # Add header with bold font.
        ws.append(HEADER)
        for cell in ws[1]:
            cell.font = bold_font

        # Select and append rows.
        stmt = (
            select(LCBMessageFilter)
            .order_by(
                LCBMessageFilter.date_of_origin,
            )
        )
        for lcb_message_filter in session.scalars(stmt):
            row = [getattr(lcb_message_filter, key) for key in HEADER]
            ws.append(row)

        # Add auto-filter
        ws.auto_filter.ref = get_used_range(ws)

        # Set column widths. These were picked manually.
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20

        wb.save(args.output)
