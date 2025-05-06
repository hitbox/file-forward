import logging

from openpyxl import Workbook
from openpyxl.styles import Font

from .base import OutputBase
from .mixin import AccumulateMixin
from .mixin import LegIdentifierMixin

logger = logging.getLogger(__name__)

class ExcelOutput(
    AccumulateMixin,
    LegIdentifierMixin,
    OutputBase,
):
    """
    Assemble source objects into an Excel report.
    """

    def __init__(
        self,
        filename,
        freeze_first_row = False,
        autofilter = False,
        filters = None,
        fields = None,
        header = None,
    ):
        """
        :param filename:
            Workbook filename to save.
        :param freeze_first_row:
            Freeze the worksheet's top (header) row.
        :param autofilter:
            Apply auto filter to worksheet.
        :param filters:
            List of pairs of (column_index, filter_values). `filter_values` can
            be a callable that returns the values. `column_index` is an index
            that the `auto_filter.add_filter_column` method can take.
        :param fields:
            Dict of keys to get from incoming file data parsed from source
            files, and callables to update worksheet cells.
        :param header:
            Optional list of callables to update cells for header row.
        """
        self.filename = filename
        self.freeze_first_row = freeze_first_row
        self.autofilter = autofilter
        self.filters = filters
        self.fields = fields
        self.header = header
        self._sources = None

    def add_autofilter(self, ws):
        """
        Add Excel auto filter to sheet. Add predefined filters if configured.
        """
        ws.auto_filter.ref = ws.dimensions
        if self.filters:
            for column_index, filter_values in self.filters:
                if callable(filter_values):
                    filter_values = filter_values(ws, column_index)
                ws.auto_filter.add_filter_column(column_index, filter_values)

    def finalize(self):
        wb = Workbook()
        ws = wb.active

        if self.header:
            for current_col, cell_writer in enumerate(self.header, start=1)
                cell = ws.cell(row=1, column=current_col)
                cell_writer(cell)
        else:
            ws.append(tuple(self.fields))

        data_rows = self.newest_by_ofp_version()
        # Write rows by one-indexed indexes. Start at two because the header is
        # already written.
        for current_row, file in enumerate(data_rows, start=ws.max_row+1):
            file_data = file.flat_dict()
            # Write cells by index, start at one; giving a callable the
            # opportunity to customize.
            for current_col, (key, cell_writer) in enumerate(self.fields.items(), start=1):
                cell = ws.cell(row=current_row, column=current_col)
                value = file_data[key]
                cell_writer(file, cell, value)

        if self.freeze_first_row:
            ws.freeze_panes = 'A2'

        if self.autofilter:
            self.add_autofilter(ws)

        wb.save(self.filename)
        logger.info('saved %s', self.filename)
